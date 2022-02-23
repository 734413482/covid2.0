import pylab as pl
import openpyxl
from sklearn.ensemble import VotingClassifier, RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import MultinomialNB
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import accuracy_score
import os
import joblib
import pandas as pd

pl.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
pl.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

data_book0 = openpyxl.load_workbook('new_普通-好转.xlsx')
data_book1 = openpyxl.load_workbook('new_普通-重症.xlsx')
data_book2 = openpyxl.load_workbook('new_重症-好转.xlsx')
data_book3 = openpyxl.load_workbook('new_重症-死亡.xlsx')

X = []
Y = []
XX_train = []
XX_test = []
row_list = []
mat_list = []

# 普通-好转
for i in range(0, len(data_book0.sheetnames)):
    data_sheet = data_book0.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(1, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(0)
    mat_list.clear()

# 普通-重症
for i in range(0, len(data_book1.sheetnames)):
    data_sheet = data_book1.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(1, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(1)
    mat_list.clear()

# 重症-好转
for i in range(0, len(data_book2.sheetnames)):
    data_sheet = data_book2.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(1, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(2)
    mat_list.clear()

# 重症-死亡
for i in range(0, len(data_book3.sheetnames)):
    data_sheet = data_book3.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(1, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(3)
    mat_list.clear()

X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.2, random_state=42)



for i in X_train:
    a1 = [y for x in i for y in x]  # 将X_train中的二维数据转化为一维
    # print(a1)
    XX_train.append(a1)

for i in X_test:
    a1 = [y for x in i for y in x]  # 将X_train中的二维数据转化为一维
    # print(a1)
    XX_test.append(a1)


print(len(X_test))
print(len(X_test[0]))
print(len(X_test[0][0]))
writer = pd.ExcelWriter('data.xlsx', engine='openpyxl')

for i in range(len(X_test)):
    test = pd.DataFrame(data=X_test[i],
                        columns=['DATE', '血_红细胞压积', '血_红细胞计数', '血_中性粒细胞(%)', '血_单核细胞(#)', '血_中性粒细胞(#)', '血_血红蛋白',
                                 '血_血小板计数', '血_白细胞计数', '血_平均血红蛋白浓度', '血_嗜碱细胞(#)', '血_单核细胞(%)', '血_嗜酸细胞(%)',
                                 '血_平均RBC体积', '血_平均血红蛋白含量', '血_淋巴细胞(#)', '血_淋巴细胞(%)', '血_嗜酸细胞(#)', '血_嗜碱细胞(%)',
                                 '血_TBIL*0.8', '血_eGFR(基于CKD-EPI方程)'])
    test.to_excel(writer, str(i), index=False)
    writer.save()

# DTC_clf = DecisionTreeClassifier(max_depth=4,
#                              min_weight_fraction_leaf=0.015,
#                              min_samples_leaf=20,
#                              min_samples_split=5,
#                              max_leaf_nodes=5,
#                              random_state=42
#                              )
# KNNuni_clf = KNeighborsClassifier(n_neighbors=9, weights="uniform", p=1)
# KNNdis_clf = KNeighborsClassifier(n_neighbors=13, weights="distance", p=1)
# MultNB_clf = MultinomialNB(alpha=0.1)
# voting_clf = VotingClassifier(
#     estimators=[('dtc', DTC_clf), ('knnu', KNNuni_clf), ('knnd', KNNdis_clf), ('multnb', MultNB_clf)],
#     voting='hard'
# )
#
# voting_clf.fit(XX_train, Y_train)
#
# for clf in (DTC_clf, KNNuni_clf, KNNdis_clf, MultNB_clf, voting_clf):
#     clf.fit(XX_train, Y_train)
#     y_pred = clf.predict(XX_test)
#     print(clf.__class__.__name__, accuracy_score(Y_test, y_pred))

