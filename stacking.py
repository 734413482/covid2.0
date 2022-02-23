"""
Purpose: This script tries to implement a technique called stacking/blending/stacked generalization.
The reason I have to make this a runnable script because I found that there isn't really any
readable code that demonstrates this technique. You may find the pseudocode in various papers but they
are all each kind of different.
Author: Eric Chio "log0" <im.ckieric@gmail.com>
======================================================================================================
Summary:
Just to test an implementation of stacking. Using a cross-validated random forest and SVMs, I was
only able to achieve an accuracy of about 88% (with 1000 trees and up). Using stacked generalization
I have seen a maximum of 93.5% accuracy. It does take runs to find it out though. This uses only
(10, 20, 10) trees for the three classifiers.
This code is heavily inspired from the code shared by Emanuele (https://github.com/emanuele) , but I
have cleaned it up to makeit available for easy download and execution.
======================================================================================================
Methodology:
Three classifiers (RandomForestClassifier, ExtraTreesClassifier and a GradientBoostingClassifier
are built to be stacked by a LogisticRegression in the end.
Some terminologies first, since everyone has their own, I'll define mine to be clear:
- DEV SET, this is to be split into the training and validation data. It will be cross-validated.
- TEST SET, this is the unseen data to validate the generalization error of our final classifier. This
set will never be used to train.
======================================================================================================
Log Output:
X_test.shape = (62L, 6L)
blend_train.shape = (247L, 3L)
blend_test.shape = (62L, 3L)
Training classifier [0]
Fold [0]
Fold [1]
Fold [2]
Fold [3]
Fold [4]
Training classifier [1]
Fold [0]
Fold [1]
Fold [2]
Fold [3]
Fold [4]
Training classifier [2]
Fold [0]
Fold [1]
Fold [2]
Fold [3]
Fold [4]
Y_dev.shape = 247
Accuracy = 0.935483870968
======================================================================================================
Data Set Information:
Biomedical data set built by Dr. Henrique da Mota during a medical residence period in the Group
of Applied Research in Orthopaedics (GARO) of the Centre MÃ©dico-Chirurgical de RÃ©adaptation des
Massues, Lyon, France. The data have been organized in two different but related classification
tasks. The first task consists in classifying patients as belonging to one out of three
categories: Normal (100 patients), Disk Hernia (60 patients) or Spondylolisthesis (150
patients). For the second task, the categories Disk Hernia and Spondylolisthesis were merged
into a single category labelled as 'abnormal'. Thus, the second task consists in classifying
patients as belonging to one out of two categories: Normal (100 patients) or Abnormal (210
patients). We provide files also for use within the WEKA environment.
Attribute Information:
Each patient is represented in the data set by six biomechanical attributes derived from the
shape and orientation of the pelvis and lumbar spine (in this order): pelvic incidence, pelvic
tilt, lumbar lordosis angle, sacral slope, pelvic radius and grade of spondylolisthesis. The
following convention is used for the class labels: DH (Disk Hernia), Spondylolisthesis (SL),
Normal (NO) and Abnormal (AB).
"""
import openpyxl
import random
import numpy as np
import matplotlib.pyplot as plt
import pylab as pl
from sklearn.model_selection import StratifiedKFold
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import LabelEncoder
from sklearn import metrics
from sklearn.naive_bayes import MultinomialNB
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.ensemble import RandomForestClassifier, ExtraTreesClassifier, GradientBoostingClassifier

pl.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
pl.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

data_book0 = openpyxl.load_workbook('new_普通-好转.xlsx')
data_book1 = openpyxl.load_workbook('new_普通-重症.xlsx')
data_book2 = openpyxl.load_workbook('new_重症-好转.xlsx')
data_book3 = openpyxl.load_workbook('new_重症-死亡.xlsx')
dataX = []
dataY = []
XX = []
row_list = []
mat_list = []

# 普通-好转
for i in range(0, len(data_book0.sheetnames)):
    data_sheet = data_book0.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(2, 21):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    dataX.append(temp_mat_list)
    dataY.append(0)
    mat_list.clear()

# 普通-重症
for i in range(0, len(data_book1.sheetnames)):
    data_sheet = data_book1.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(2, 21):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    dataX.append(temp_mat_list)
    dataY.append(1)
    mat_list.clear()

# # 重症-好转
# for i in range(0, len(data_book2.sheetnames)):
#     data_sheet = data_book2.worksheets[i]
#     data_row = data_sheet.max_row
#     for j in range(2, data_row+1):
#         for k in range(2, 21):
#             row_list.append(data_sheet.cell(row=j, column=k).value)
#         temp_row_list = list(row_list)
#         mat_list.append(temp_row_list)
#         row_list.clear()
#     temp_mat_list = list(mat_list)
#     dataX.append(temp_mat_list)
#     dataY.append(2)
#     mat_list.clear()
#
# # 重症-死亡
# for i in range(0, len(data_book3.sheetnames)):
#     data_sheet = data_book3.worksheets[i]
#     data_row = data_sheet.max_row
#     for j in range(2, data_row+1):
#         for k in range(2, 21):
#             row_list.append(data_sheet.cell(row=j, column=k).value)
#         temp_row_list = list(row_list)
#         mat_list.append(temp_row_list)
#         row_list.clear()
#     temp_mat_list = list(mat_list)
#     dataX.append(temp_mat_list)
#     dataY.append(3)
#     mat_list.clear()

for i in dataX:
    a1 = [y for x in i for y in x]  # 将X_train中的二维数据转化为一维
    # print(a1)
    XX.append(a1)

def run(X,Y):
    # We need to transform the string output to numeric
    label_encoder = LabelEncoder()
    label_encoder.fit(Y)
    Y = label_encoder.transform(Y)

    # The DEV SET will be used for all training and validation purposes
    # The TEST SET will never be used for training, it is the unseen set.
    dev_cutoff = len(Y) * 4 / 5
    dev_cutoff = int(dev_cutoff)
    X_dev = X[:dev_cutoff]
    Y_dev = Y[:dev_cutoff]
    X_test = X[dev_cutoff:]
    Y_test = Y[dev_cutoff:]

    n_folds = 5
    # Our level 0 classifiers
    clfs = [
        DecisionTreeClassifier(max_depth=4,
                               min_weight_fraction_leaf=0.015,
                               min_samples_leaf=20,
                               min_samples_split=5,
                               max_leaf_nodes=5,
                               random_state=42
                               ),
        KNeighborsClassifier(n_neighbors=9, weights="uniform", p=1),
        KNeighborsClassifier(n_neighbors=13, weights="distance", p=1),
        MultinomialNB(alpha=0.1)
    ]

    # Ready for cross validation
    skf = list(StratifiedKFold(Y_dev, int(n_folds)))

    # Pre-allocate the data
    blend_train = np.zeros((X_dev.shape[0], len(clfs)))  # Number of training data x Number of classifiers
    blend_test = np.zeros((X_test.shape[0], len(clfs)))  # Number of testing data x Number of classifiers

    print
    'X_test.shape = %s' % (str(X_test.shape))
    print
    'blend_train.shape = %s' % (str(blend_train.shape))
    print
    'blend_test.shape = %s' % (str(blend_test.shape))

    # For each classifier, we train the number of fold times (=len(skf))
    for j, clf in enumerate(clfs):
        print
        'Training classifier [%s]' % (j)
        blend_test_j = np.zeros((X_test.shape[0], len(skf)))  # Number of testing data x Number of folds , we will take the mean of the predictions later
        for i, (train_index, cv_index) in enumerate(skf):
            print
            'Fold [%s]' % (i)

            # This is the training and validation set
            X_train = X_dev[train_index]
            Y_train = Y_dev[train_index]
            X_cv = X_dev[cv_index]
            Y_cv = Y_dev[cv_index]

            clf.fit(X_train, Y_train)

            # This output will be the basis for our blended classifier to train against,
            # which is also the output of our classifiers
            blend_train[cv_index, j] = clf.predict(X_cv)
            blend_test_j[:, i] = clf.predict(X_test)
        # Take the mean of the predictions of the cross validation set
        blend_test[:, j] = blend_test_j.mean(1)

    print
    'Y_dev.shape = %s' % (Y_dev.shape)

    # Start blending!
    bclf = LogisticRegression()
    bclf.fit(blend_train, Y_dev)

    # Predict now
    Y_test_predict = bclf.predict(blend_test)
    score = metrics.accuracy_score(Y_test, Y_test_predict)
    print
    'Accuracy = %s' % (score)

    return score


if __name__ == '__main__':
    train_file = 'data/column_3C.dat'

    data = XX
    best_score = 0.0

    # run many times to get a better result, it's not quite stable.
    for i in range(1):
        print
        'Iteration [%s]' % (i)
        random.shuffle(data)
        dataY = np.array(dataY)
        data = np.array(data)
        score = run(data,dataY)
        best_score = max(best_score, score)
        print

    print
    'Best score = %s' % (best_score)