import openpyxl
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import MultinomialNB
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import accuracy_score, classification_report
from sklearn.ensemble import VotingClassifier, RandomForestClassifier
from sklearn import preprocessing
import numpy as np
import matplotlib.pyplot as plt
import pylab as pl

data_book = openpyxl.load_workbook('ct名单new.xlsx')
data_sheet = data_book.worksheets[1]
X = []
Y = []

for i in range (int((data_sheet.max_row)/3)):
    j = i*3 + 1
    Y.append(data_sheet.cell(row=j, column=3).value)
    data_list = []
    data_list.append(data_sheet.cell(row=j, column=4).value / data_sheet.cell(row=j, column=5).value)
    data_list.append(data_sheet.cell(row=j+1, column=4).value / data_sheet.cell(row=j+1, column=5).value)
    data_list.append(data_sheet.cell(row=j+2, column=4).value / data_sheet.cell(row=j+2, column=5).value)
    X.append(data_list)

print(max(X))

X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.2, random_state=42)
print(X_test)