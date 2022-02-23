import numpy as np
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import KFold, train_test_split
import openpyxl
import pylab as pl

pl.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
pl.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

data_book0 = openpyxl.load_workbook('new_普通-好转.xlsx')
data_book1 = openpyxl.load_workbook('new_普通-重症.xlsx')
data_book2 = openpyxl.load_workbook('new_重症-好转.xlsx')
data_book3 = openpyxl.load_workbook('new_重症-死亡.xlsx')
X = []
Y = []
XX_train = []
XX_test = []
row_list = []
mat_list = []

# 普通-好转
for i in range(0, len(data_book0.sheetnames)):
    data_sheet = data_book0.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(2, 21):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(0)
    mat_list.clear()

# 普通-重症
for i in range(0, len(data_book1.sheetnames)):
    data_sheet = data_book1.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(2, 21):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(1)
    mat_list.clear()

# 重症-好转
for i in range(0, len(data_book2.sheetnames)):
    data_sheet = data_book2.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(2, 21):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(2)
    mat_list.clear()

# 重症-死亡
for i in range(0, len(data_book3.sheetnames)):
    data_sheet = data_book3.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(2, 21):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(3)
    mat_list.clear()

def get_stacking(clf, x_train, y_train, x_test, n_folds=10):
    """
    这个函数是stacking的核心，使用交叉验证的方法得到次级训练集
    x_train, y_train, x_test 的值应该为numpy里面的数组类型 numpy.ndarray .
    如果输入为pandas的DataFrame类型则会把报错"""
    train_num, test_num = x_train.shape[0], x_test.shape[0]
    second_level_train_set = np.zeros((train_num,))
    second_level_test_set = np.zeros((test_num,))
    test_nfolds_sets = np.zeros((test_num, n_folds))
    kf = KFold(n_splits=n_folds)
    print(kf.split(x_train))
    for i,(train_index, test_index) in enumerate(kf.split(x_train)):
        x_tra, y_tra = x_train[train_index], y_train[train_index]
        x_tst, y_tst =  x_train[test_index], y_train[test_index]

        clf.fit(x_tra, y_tra)

        second_level_train_set[test_index] = clf.predict(x_tst)
        test_nfolds_sets[:,i] = clf.predict(x_test)

    second_level_test_set[:] = test_nfolds_sets.mean(axis=1)
    return second_level_train_set, second_level_test_set



from sklearn.naive_bayes import MultinomialNB
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier

DTC_clf = DecisionTreeClassifier(max_depth=4,
                             min_weight_fraction_leaf=0.015,
                             min_samples_leaf=20,
                             min_samples_split=5,
                             max_leaf_nodes=5,
                             random_state=42
                             )
KNNuni_clf = KNeighborsClassifier(n_neighbors=9, weights="uniform", p=1)
KNNdis_clf = KNeighborsClassifier(n_neighbors=13, weights="distance", p=1)
MultNB_clf = MultinomialNB(alpha=0.1)

X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.2, random_state=42)

for i in X_train:
    a1 = [y for x in i for y in x]  # 将X_train中的二维数据转化为一维
    # print(a1)
    XX_train.append(a1)

for i in X_test:
    a1 = [y for x in i for y in x]  # 将X_train中的二维数据转化为一维
    # print(a1)
    XX_test.append(a1)
XX_train = np.array(XX_train)
XX_test = np.array(XX_test)
Y_train = np.array(Y_train)
train_sets = []
test_sets = []
for clf in [DTC_clf, KNNuni_clf, KNNdis_clf, MultNB_clf]:
    train_set, test_set = get_stacking(clf, XX_train, Y_train, XX_test)
    print(train_set)
    train_sets.append(train_set)
    test_sets.append(test_set)

meta_train = np.concatenate([result_set.reshape(-1,1) for result_set in train_sets], axis=1)
meta_test = np.concatenate([y_test_set.reshape(-1,1) for y_test_set in test_sets], axis=1)
print(meta_train)

dt_model = LogisticRegression()
dt_model.fit(meta_train, Y_train)
df_predict = dt_model.predict(meta_test)

print(df_predict)
print(Y_test)
right = 0
for i in range(len(df_predict)):
    if(df_predict[i]==Y_test[i]):
        right = right + 1
print(right/35)