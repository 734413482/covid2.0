import openpyxl

data_book0 = openpyxl.load_workbook('1filleddelete_普通-好转.xlsx')
data_book1 = openpyxl.load_workbook('1filleddelete_普通-重症.xlsx')
data_book2 = openpyxl.load_workbook('1filleddelete_重症-好转.xlsx')
data_book3 = openpyxl.load_workbook('1filleddelete_重症-死亡.xlsx')

book0_name = []
book1_name = []
book2_name = []
book3_name = []
book0_list = []
book1_list = []
book2_list = []
book3_list = []
for i in range(0, len(data_book0.sheetnames)):
    book0_name.append(data_book0.sheetnames[i])

for i in range(0, len(data_book1.sheetnames)):
    book1_name.append(data_book1.sheetnames[i])

for i in range(0, len(data_book2.sheetnames)):
    book2_name.append(data_book2.sheetnames[i])

for i in range(0, len(data_book3.sheetnames)):
    book3_name.append(data_book3.sheetnames[i])

for i in range(0,len(book0_name)):
    index = 0
    for j in range(0,len(book0_name[i])):
        if(book0_name[i][j]=='-'):
            index = j+1
    book0_list.append(book0_name[i][index:index+10])
for i in range(0,len(book1_name)):
    index = 0
    for j in range(0,len(book1_name[i])):
        if(book1_name[i][j]=='-'):
            index = j+1
    book1_list.append(book1_name[i][index:index+10])
for i in range(0,len(book2_name)):
    index = 0
    for j in range(0,len(book2_name[i])):
        if(book2_name[i][j]=='-'):
            index = j+1
    book2_list.append(book2_name[i][index:index+10])
for i in range(0,len(book3_name)):
    index = 0
    for j in range(0,len(book3_name[i])):
        if(book3_name[i][j]=='-'):
            index = j+1
    book3_list.append(book3_name[i][index:index+10])

repeat = []

for i in range (0,len(book2_list)):
    for j in range(0,len(book3_list)):
        if(book2_list[i]==book3_list[j]):
            repeat.append(book2_list[i])

print(repeat)