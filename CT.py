import openpyxl
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.model_selection import train_test_split
from sklearn.naive_bayes import MultinomialNB
from sklearn.tree import DecisionTreeClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import accuracy_score, classification_report
from sklearn.ensemble import VotingClassifier, RandomForestClassifier
from sklearn import preprocessing
import numpy as np
import matplotlib.pyplot as plt
import pylab as pl

data_book = openpyxl.load_workbook('ct名单new.xlsx')
data_sheet = data_book.worksheets[1]
X = []
Y = []

for i in range (int((data_sheet.max_row)/3)):
    j = i*3 + 1
    Y.append(data_sheet.cell(row=j, column=3).value)
    data_list = []
    data_list.append(data_sheet.cell(row=j, column=4).value / data_sheet.cell(row=j, column=5).value)
    data_list.append(data_sheet.cell(row=j+1, column=4).value / data_sheet.cell(row=j+1, column=5).value)
    data_list.append(data_sheet.cell(row=j+2, column=4).value / data_sheet.cell(row=j+2, column=5).value)
    X.append(data_list)

X = np.array(X)
min_max_scaler = preprocessing.MinMaxScaler()
X = min_max_scaler.fit_transform(X)

X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.2, random_state=42)
print(X_test)
KNNuni_clf = KNeighborsClassifier(n_neighbors=7, weights="uniform", p=1)

KNNuni_clf.fit(X_train, Y_train)
y_pred = KNNuni_clf.predict(X_test)
print(KNNuni_clf.__class__.__name__, accuracy_score(Y_test, y_pred))
print(y_pred)
print(KNNuni_clf.predict_proba(X_test))

knn = KNeighborsClassifier(n_neighbors=7, weights="uniform", p=1)
scores = cross_val_score(knn,X,Y,cv=10,scoring='accuracy')
print(scores)