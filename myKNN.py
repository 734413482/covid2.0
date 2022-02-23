'''
k近邻（kNN）算法的工作机制比较简单，根据某种距离测度找出距离给定待测样本距离最小的k个训练样本，根据k个训练样本进行预测。
分类问题：k个点中出现频率最高的类别作为待测样本的类别
回归问题：通常以k个训练样本的平均值作为待测样本的预测值
kNN模型三要素：距离测度、k值的选择、分类或回归决策方式
'''
import numpy as np
import openpyxl
from matplotlib import pylab as pl
from sklearn.model_selection import train_test_split
from sklearn.neighbors import KNeighborsClassifier

pl.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
pl.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

data_book0 = openpyxl.load_workbook('new_普通-好转.xlsx')
data_book1 = openpyxl.load_workbook('new_普通-重症.xlsx')
data_book2 = openpyxl.load_workbook('new_重症-好转.xlsx')
data_book3 = openpyxl.load_workbook('new_重症-死亡.xlsx')

X = []
Y = []
XX_train = []
XX_test = []
row_list = []
mat_list = []

# 普通-好转
for i in range(0, len(data_book0.sheetnames)):
    data_sheet = data_book0.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(1, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(0)
    mat_list.clear()
# print(X)
# print(Y)

# 普通-重症
for i in range(0, len(data_book1.sheetnames)):
    data_sheet = data_book1.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(1, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(1)
    mat_list.clear()

# 重症-好转
for i in range(0, len(data_book2.sheetnames)):
    data_sheet = data_book2.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(1, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(2)
    mat_list.clear()

# 重症-死亡
for i in range(0, len(data_book3.sheetnames)):
    data_sheet = data_book3.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(1, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(3)
    mat_list.clear()

X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.2, random_state=42)

class KNNClassfier(object):

    def __init__(self, k=9, distance='euc'):
        self.k = k
        self.distance = distance
        self.x = None
        self.y = None
    def fit(self,X, Y):
        '''
        X : array-like [n_samples,shape]
        Y : array-like [n_samples,1]
        '''
        self.x = X
        self.y = Y
    def predict(self,X_test):
        '''
        X_test : array-like [n_samples,shape]
        Y_test : array-like [n_samples,1]
        output : array-like [n_samples,1]
        '''
        output = np.zeros((X_test.shape[0],1))
        for i in range(X_test.shape[0]):
            dis = []
            for j in range(self.x.shape[0]):
                if self.distance == 'euc': # 欧式距离
                    #dis.append(np.linalg.norm(X_test[i]-self.x[j,:],ord=2, axis=None, keepdims=True))  #k=17
                    dis.append(abs(np.linalg.norm(X_test[i],ord=2, axis=None, keepdims=False)-
                                   np.linalg.norm(self.x[j,:],ord=2, axis=None, keepdims=False)))      #k=9
            labels = []
            index=sorted(range(len(dis)), key=dis.__getitem__)
            for j in range(self.k):
                labels.append(self.y[index[j]])
            counts = []
            for label in labels:
                counts.append(labels.count(label))
            output[i] = labels[np.argmax(counts)]
        return output
    def score(self,x,y):
        pred = self.predict(x)
        err = 0.0
        for i in range(x.shape[0]):
            if pred[i]!=y[i]:
                err = err+1
        return 1-float(err/x.shape[0])


if __name__ == '__main__':
    from sklearn import datasets
    iris = datasets.load_iris()
    x = np.array(X_train)
    y = np.array(Y_train)
    # print(x.shape[0])
    # x = np.array([[0.5,0.4],[0.1,0.2],[0.7,0.8],[0.2,0.1],[0.4,0.6],[0.9,0.9],[1,1]]).reshape(-1,2)
    # y = np.array([0,1,0,1,0,1,1]).reshape(-1,1)
    clf = KNNClassfier(k=9)
    clf.fit(x,y)
    print('myknn train score:', clf.score(np.array(X_train), np.array(Y_train)))
    print('myknn test score:',clf.score(np.array(X_test),np.array(Y_test)))
    # print(x[0])
    # print(x[1])
    print(np.linalg.norm(x[0]-x[1]))