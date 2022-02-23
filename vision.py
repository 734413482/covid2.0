import numpy as np
from numpy import *
from scipy import interpolate
import pylab as pl
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

from sklearn.ensemble import ExtraTreesClassifier, RandomForestClassifier
from sklearn.naive_bayes import BernoulliNB, MultinomialNB
from sklearn.naive_bayes import GaussianNB
from sklearn.model_selection import train_test_split
from sklearn.tree import DecisionTreeClassifier
from sklearn import tree, metrics

from sklearn.neighbors import KNeighborsClassifier
from sklearn.metrics import confusion_matrix, classification_report
pl.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签
pl.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号

data_book0 = openpyxl.load_workbook('1filled_普通-好转.xlsx')
data_book1 = openpyxl.load_workbook('1filled_普通-重症.xlsx')
data_book2 = openpyxl.load_workbook('1filled_重症-好转.xlsx')
data_book3 = openpyxl.load_workbook('1filled_重症-死亡.xlsx')
data_book4 = openpyxl.load_workbook('testonedata.xlsx')
X = []
Y = []
XX_train = []
XX_test = []
row_list = []
mat_list = []
my_test = []
my_tests = []
classname = []

data_sheet = data_book0.worksheets[0]
for i in range(2,22):
    classname.append(data_sheet.cell(row=1, column=i).value)

for i in range(0, len(data_book4.sheetnames)):
    data_sheet = data_book4.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    my_test.append(temp_mat_list)
    mat_list.clear()
for i in my_test:
    a1 = [y for x in i for y in x]  # 将X_train中的二维数据转化为一维
    # print(a1)
    my_tests.append(a1)

# 普通-好转
for i in range(0, len(data_book0.sheetnames)):
    data_sheet = data_book0.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(0)
    mat_list.clear()

# 普通-重症
for i in range(0, len(data_book1.sheetnames)):
    data_sheet = data_book1.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(1)
    mat_list.clear()

# 重症-好转
for i in range(0, len(data_book2.sheetnames)):
    data_sheet = data_book2.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(2)
    mat_list.clear()

# 重症-死亡
for i in range(0, len(data_book3.sheetnames)):
    data_sheet = data_book3.worksheets[i]
    data_row = data_sheet.max_row
    for j in range(2, data_row+1):
        for k in range(2, 22):
            row_list.append(data_sheet.cell(row=j, column=k).value)
        temp_row_list = list(row_list)
        mat_list.append(temp_row_list)
        row_list.clear()
    temp_mat_list = list(mat_list)
    X.append(temp_mat_list)
    Y.append(3)
    mat_list.clear()

X_train, X_test, Y_train, Y_test = train_test_split(X, Y, test_size=0.2, random_state=42)

for i in X_train:
    a1 = [y for x in i for y in x]  # 将X_train中的二维数据转化为一维
    # print(a1)
    XX_train.append(a1)

for i in X_test:
    a1 = [y for x in i for y in x]  # 将X_train中的二维数据转化为一维
    # print(a1)
    XX_test.append(a1)

def test_BernoulliNB_alpha(*data):
    '''
    测试 BernoulliNB 的预测性能随 alpha 参数的影响
    '''
    X_train,X_test,y_train,y_test=data
    alphas=np.logspace(-2,5,num=200)
    train_scores=[]
    test_scores=[]
    for alpha in alphas:
        cls=MultinomialNB(alpha=alpha)
        cls.fit(X_train,y_train)
        train_scores.append(cls.score(X_train,y_train))
        test_scores.append(cls.score(X_test, y_test))

    ## 绘图
    fig=pl.plt.figure()
    ax=fig.add_subplot(1,1,1)
    ax.plot(alphas,train_scores,label="Training Score")
    ax.plot(alphas,test_scores,label="Testing Score")
    ax.set_xlabel(r"$\alpha$")
    ax.set_ylabel("score")
    ax.set_ylim(0,1.0)
    ax.set_title("BernoulliNB")
    ax.set_xscale("log")
    ax.legend(loc="best")
    pl.plt.show()
#test_BernoulliNB_alpha(XX_train,XX_test,Y_train,Y_test)

# 准确率与权重和k值关系
def test_KNeighborsClassifier_k_w(*data):
    X_train, X_test, Y_train, Y_test = data
    Ks = np.linspace(1, len(Y_train), num=100, endpoint=False, dtype="int")
    # uniform：投票权重相同
    # distance:投票权重跟距离成反比
    weights = ["uniform", "distance"]

    fig = plt.figure()
    ax = fig.add_subplot(1, 1, 1)
    for weight in weights:
        training_scores = []
        testing_scores = []
        for K in Ks:
            clf = KNeighborsClassifier(weights=weight, n_neighbors=K, p=1)
            clf.fit(X_train, Y_train)
            training_scores.append(clf.score(X_train, Y_train))
            testing_scores.append(clf.score(X_test, Y_test))
        ax.plot(Ks, training_scores, label="Training score:weight=%s" % weight)
        ax.plot(Ks, testing_scores, label="Testing score:weight=%s" % weight)
    ax.legend(loc="best")
    ax.set_xlabel("K")
    ax.set_ylabel("score")
    ax.set_ylim(0, 1.05)
    ax.set_title("KNeighborsClssifier")
    plt.show()
#test_KNeighborsClassifier_k_w(XX_train, XX_test, Y_train, Y_test)
maxLeaves = range(2, 20)
score_train=np.zeros(20)
score_test=np.zeros(20)
for maxLeaf in maxLeaves:
    clf = DecisionTreeClassifier(max_depth=3,
                                min_weight_fraction_leaf=0.015,
                                min_samples_leaf=20,
                                min_samples_split=5,
                                #max_features=17,
                                max_leaf_nodes=maxLeaf,
                                random_state=42)
    clf.fit(XX_train, Y_train)

    print("train: ", clf.score(XX_train, Y_train))
    print("test: ", clf.score(XX_test, Y_test))

    score_test[maxLeaf] = clf.score(XX_test,Y_test)
    score_train[maxLeaf] = clf.score(XX_train,Y_train)

plt.figure(figsize=(10,6))
sns.set(style="whitegrid")
data = pd.DataFrame({"score_train": score_train, "score_test": score_test})
sns.lineplot(data=data)
plt.xlabel("max_leaf_nodes")
plt.ylabel("accuracy")
plt.title("accuracy varies with max_leaf_nodes")
plt.show()